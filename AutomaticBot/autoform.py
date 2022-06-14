import webbrowser
from openpyxl import load_workbook
import pyautogui as pg
import pyperclip
import random
import time
import wikipedia
import os

url = 'http://uncle-machine.com/'
webbrowser.open(url)
time.sleep(2)
pg.hotkey('win', 'up') # Maximize Screen
time.sleep(2)

# กดปุ่มตามรูป
'''
path = r'D:\AutomaticBot'
bt1 = os.path.join(path, 'sell_button.png')
pg.click(bt1)
#pg.click('sell_button.png')
'''

pg.click(200,200)

pg.press('tab', presses=3)
pg.hotkey('enter')

excelfile = load_workbook('fruit.xlsx')
sheet = excelfile.active
#print(sheet['B2'].value)

count = len(sheet['B'])
#print(count)

fruit_list = []

for i in range(2,count+1):
    data = sheet.cell(row=i,column=2).value
    #print(data)
    split = data.split(',')
    if len(split) >= 2: #เเบ่งคำ เช่น Romaine Lettuce, chopped กลายเป็น ['Romaine Lettuce', 'chopped']
        print(split[0])
        if split[0] not in fruit_list:
            fruit_list.append(split[0])
    else:
        split = data.split('(')
        print(split)
        if split[0] not in fruit_list:
            fruit_list.append(split[0])         

    fruit_list.append(data)

print(fruit_list)
print('==========================')

#*********  Work Flow *********

# เผื่อกรณีขี้เกียจคลิ๊ก tab google สลับกับ youtube
#pg.click(1184,13)

# คลิ๊กที่เว็บ browser
pg.click(200,200)

for f in fruit_list:

    # กด Tab 6 ครังเพื่อให้ Cursor อยู่ในช่องกรอก
    pg.press('tab', presses=6)     

    # กรอกข้อมูลชื่อ + Tab
    product = f
    pyperclip.copy(product)
    pg.hotkey('ctrl','v')
    pg.press('tab')

    # สุ่มราคา
    rand = random.choice(range(100,1001,100))

    # กรอกราคา (ตัวเลข) + Tab
    pyperclip.copy(rand)
    pg.hotkey('ctrl','v')
    pg.press('tab')

    # ข้อมูลสินค้า กรอกชื่อร้านไปก่อน + Tab
    try:
        info = wikipedia.summary(f)
    except:
        info = 'No Detail - Ticky Fruit Shop'
    pyperclip.copy(info)
    pg.hotkey('ctrl','v')
    pg.press('tab')

    # Tab เพื่อไปยังปุ่ม submit
    pg.press('tab')

    # Enter เพื่อ submit
    pg.press('enter')
    #time.sleep(1)    